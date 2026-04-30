from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.contrib.auth import authenticate
from django.contrib.auth import get_user_model
from .serializers import UsuarioSerializer, UsuarioGymSerializer, UsuarioGymDaySerializer, MembresiasSerializer, MembresiaAsignadaSerializer
from .models import Usuario, UsuarioGym, UsuarioGymDay, Membresia, MembresiaAsignada, Gimnasio
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
#para la imagen
from rest_framework.parsers import MultiPartParser, FormParser
#Para las filtraciones en la base de datos
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter
#Para las notificaciones
from datetime import datetime, date, timedelta
from rest_framework.decorators import api_view, permission_classes
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.conf import settings


# Importar permisos personalizados
from .permissions import IsAdminUser, IsRecepcionUser
from .mixins import MultiTenantViewSetMixin


# ============================================================
# VIEWSETS
# ============================================================

class UserViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]  # Solo admins pueden gestionar usuarios
    parser_classes = (MultiPartParser, FormParser)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                "error": "Datos inválidos",
                "details": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        self.perform_create(serializer)
        user_data = serializer.instance

        return Response({
            "user": UsuarioSerializer(user_data, context={'request': request}).data,
        }, status=status.HTTP_201_CREATED)
    
    def update(self, request, *args, **kwargs):
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()

            old_avatar = instance.avatar if instance.avatar else None

            serializer = self.get_serializer(
                instance, 
                data=request.data, 
                partial=partial,
                context={'request': request}
            )

            if serializer.is_valid():
                if 'avatar' in request.FILES:
                    if old_avatar:
                        old_avatar.delete(save=False)
                    instance.avatar = request.FILES['avatar']

                self.perform_update(serializer)
                return Response(serializer.data, status=status.HTTP_200_OK)
            
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({"error": "Error inesperado"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Registro público para crear usuario inicial
# El login se maneja con SimpleJWT en /token/ (TokenObtainPairView)
class RegisterViewSet(APIView):
    permission_classes = [AllowAny]
    
    def post(self, request, *args, **kwargs):
        from rest_framework_simplejwt.tokens import RefreshToken
        
        email = request.data.get('email')
        password = request.data.get('password')
        name = request.data.get('name')
        lastname = request.data.get('lastname')
        
        if not email or not password or not name or not lastname:
            return Response({
                'error': 'Todos los campos son requeridos',
                'required': ['email', 'password', 'name', 'lastname']
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Verificar si el usuario ya existe
        if get_user_model().objects.filter(email=email).exists():
            return Response({'error': 'El correo ya está registrado'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Crear gimnasio automáticamente
        email_prefix = email.split('@')[0].replace('.', ' ').title()
        gimnasio = Gimnasio.objects.create(name=f"Gimnasio {email_prefix}")
        
        # Crear usuario con gimnasio
        user = get_user_model()(
            email=email,
            name=name,
            lastname=lastname,
            roles='admin',  # Primer usuario es admin
            gimnasio=gimnasio
        )
        user.set_password(password)
        user.save()
        
        # Generar tokens JWT
        refresh = RefreshToken.for_user(user)
        
        response = Response({
            'message': 'Usuario creado exitosamente',
            'user': UsuarioSerializer(user, context={'request': request}).data,
            'access': str(refresh.access_token),
        }, status=status.HTTP_201_CREATED)
        
        # Establecer refresh token como cookie HttpOnly
        response.set_cookie(
            key='refresh_token',
            value=str(refresh),
            max_age=604800,  # 7 days
            httponly=True,
            secure=not settings.DEBUG,
            samesite='Lax',
            path='/gym/api/v1/token/refresh/',
        )
        
        return response


class userProfileView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            user_serializer = UsuarioSerializer(request.user, context={'request': request})
            user_data = user_serializer.data            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'user': user_data}, status=status.HTTP_200_OK)
          
    def list(self, request):
        try:
            gimnasio = request.gimnasio
            if gimnasio:
                users = Usuario.objects.filter(gimnasio=gimnasio).order_by('-id')
            else:
                users = Usuario.objects.none()
            serializer = UsuarioSerializer(users, many=True, context={'request': request})
            return Response({'users': serializer.data}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)    


# ============================================================
# MIEMBROS DEL GIMNASIO
# ============================================================

class UsuarioGymViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    queryset = UsuarioGym.objects.all()
    serializer_class = UsuarioGymSerializer
    permission_classes = [IsAuthenticated, IsRecepcionUser]  # Admin y recepcion pueden acceder
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['name', 'lastname']


class UsuarioGymDayViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    queryset = UsuarioGymDay.objects.all()
    serializer_class = UsuarioGymDaySerializer
    permission_classes = [IsAuthenticated, IsRecepcionUser]  # Admin y recepcion pueden acceder
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['name', 'lastname']

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                "error": "Datos inválidos para ingreso diario",
                "details": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


# ============================================================
# ESTADÍSTICAS DEL DASHBOARD (Miembros Activos, Nuevos, Retention)
# ============================================================

class DashboardStatsView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        from django.db.models import Count, Q
        from datetime import date, timedelta
        
        gimnasio = request.gimnasio
        
        if not gimnasio:
            return Response({
                'active_members': 0,
                'new_today': 0,
                'retention_rate': 0
            })
        
        today = date.today()
        
        # 1. Miembros activos (membresías vigentes)
        active_memberships = MembresiaAsignada.objects.filter(
            miembro__gimnasio=gimnasio,
            dateInitial__lte=today,
            dateFinal__gte=today
        )
        active_count = active_memberships.count()
        
        # 2. Nuevos hoy (miembros que se registraronHOY)
        new_today = active_memberships.filter(
            dateInitial=today
        ).count()
        
        # 3. Retention Rate
        # Mes anterior
        if today.month == 1:
            prev_month = 12
            prev_year = today.year - 1
        else:
            prev_month = today.month - 1
            prev_year = today.year
        
        # Primer día del mes anterior
        prev_month_start = date(prev_year, prev_month, 1)
        
        # Último día del mes anterior
        if prev_month == 12:
            prev_month_end = date(prev_year + 1, 1, 1) - timedelta(days=1)
        else:
            prev_month_end = date(prev_year, prev_month + 1, 1) - timedelta(days=1)
        
        # Miembros que tenían membresía activa el mes anterior
        prev_active = MembresiaAsignada.objects.filter(
            miembro__gimnasio=gimnasio,
            dateInitial__lte=prev_month_end,
            dateFinal__gte=prev_month_start
        ).count()
        
        # Miembros que estavam activos el mes anterior Y aún siguen activos
        from django.db.models import Q
        still_active = MembresiaAsignada.objects.filter(
            miembro__gimnasio=gimnasio,
            dateInitial__lte=prev_month_end,
            dateFinal__gte=prev_month_start
        ).filter(
            dateFinal__gte=today
        ).count()
        
        # Calcular retention
        if prev_active > 0:
            retention = round((still_active / prev_active) * 100, 1)
        else:
            retention = 100.0  # Si no hay miembros anteriores, 100%
        
        return Response({
            'active_members': active_count,
            'new_today': new_today,
            'retention_rate': retention
        })


# ============================================================
# HOME / DASHBOARD
# ============================================================

class Home(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        gimnasio = request.gimnasio
        
        if gimnasio:
            UserGymList = MembresiaAsignada.objects.filter(miembro__gimnasio=gimnasio).order_by('-id')
            UserDayList = UsuarioGymDay.objects.filter(gimnasio=gimnasio).order_by('-id')
        else:
            UserGymList = MembresiaAsignada.objects.none()
            UserDayList = UsuarioGymDay.objects.none()

        now = timezone.now()
        month = now.month
        year = now.year

        # Mes anterior
        if month == 1:
            prev_month, prev_year = 12, year - 1
        else:
            prev_month, prev_year = month - 1, year

        num_miembros = UserGymList.count()

        miembros_mes = UserGymList.filter(dateInitial__month=month, dateInitial__year=year)
        total_gym_mes = sum(user.price for user in miembros_mes)

        miembrosDay_mes = UserDayList.filter(dateInitial__month=month, dateInitial__year=year)
        total_day_mes = sum(user.price for user in miembrosDay_mes)

        total_month = total_gym_mes + total_day_mes
        miembros_mes_count = miembros_mes.count()

        # Miembros del mes anterior
        miembros_mes_anterior = UserGymList.filter(
            dateInitial__month=prev_month,
            dateInitial__year=prev_year
        ).count()

        # Diferencia vs mes anterior
        diff_miembros = miembros_mes_count - miembros_mes_anterior

        total_gym = sum(user.price for user in UserGymList)
        total_day = sum(user.price for user in UserDayList)
        total = total_gym + total_day

        return JsonResponse({ 
            'num_miembros': num_miembros, 
            'total_month': float(total_month), 
            'miembros_mes': miembros_mes_count,
            'total': float(total),
            'miembros_mes_anterior': miembros_mes_anterior,   
            'diff_miembros': diff_miembros,                   
        })
    

# ============================================================
# MEMBRESIAS
# ============================================================

class MembresiaViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    queryset = Membresia.objects.all()
    serializer_class = MembresiasSerializer
    permission_classes = [IsAuthenticated, IsRecepcionUser]


class MembresiaAsignadaViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    queryset = MembresiaAsignada.objects.all()
    serializer_class = MembresiaAsignadaSerializer
    permission_classes = [IsAuthenticated, IsRecepcionUser]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['miembro__name', 'miembro__lastname']
    gimnasio_field = 'miembro__gimnasio'

    def get_queryset(self):
        """Filtrar membresías asignadas por gimnasio del usuario actual."""
        queryset = super().get_queryset()
        
        # Filtro adicional por miembro si se pasa
        miembro_id = self.request.query_params.get('miembro')
        if miembro_id:
            queryset = queryset.filter(miembro_id=miembro_id)
        
        return queryset.order_by('-id')


# ============================================================
# NOTIFICATIONS
# ============================================================

# ============================================================
# ACTIVIDADES RECIENTES
# ============================================================

class ActivitiesView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        gimnasio = request.gimnasio
        
        activities = []
        
        if gimnasio:
            membresias_recientes = MembresiaAsignada.objects.filter(
                miembro__gimnasio=gimnasio
            ).select_related('miembro', 'membresia').order_by('-created_at')[:5]

            ingresos = UsuarioGymDay.objects.filter(
                gimnasio=gimnasio
            ).order_by('-created_at')[:5]
            
            for membresia in membresias_recientes:
                
                activities.append({
                    'id': f'm-{membresia.id}',
                    'type': 'new_member',
                    'icon': 'person_add',
                    'color': 'primary',
                    'title': 'Nuevo miembro registrado',
                    'description': f'{membresia.miembro.name} {membresia.miembro.lastname} - {membresia.membresia.name}',                    
                    'created_at': membresia.created_at.isoformat(),
                    'time_ago': self.get_time_ago(membresia.created_at),
                })            
            
            
            for ingreso in ingresos:
                
                activities.append({
                    'id': f'i-{ingreso.id}',
                    'type': 'entry',
                    'icon': 'login',
                    'color': 'info',
                    'title': 'Ingreso registrado',
                    'description': f'{ingreso.name} {ingreso.lastname}',
                    'created_at': ingreso.created_at.isoformat(),
                    'amount': float(ingreso.price),
                    'time_ago': self.get_time_ago(ingreso.created_at)
                })
        
        activities.sort(key=lambda x: x['created_at'], reverse=True)
        
        return Response(activities[:10])
    
     # ==========================================================
    # HELPERS
    # ==========================================================

    def ensure_datetime(self, date_obj):
        """Convierte date → datetime seguro"""
        if isinstance(date_obj, date) and not isinstance(date_obj, datetime):
            date_obj = datetime.combine(date_obj, datetime.min.time())

        if timezone.is_naive(date_obj):
            date_obj = timezone.make_aware(date_obj)

        return date_obj
    
    def get_time_ago(self, date_obj):
        now = timezone.now()
        diff = now - date_obj

        if diff.days > 0:
            return f'{diff.days} día{"s" if diff.days != 1 else ""}'
        elif diff.seconds >= 3600:
            hours = diff.seconds // 3600
            return f'{hours} hora{"s" if hours != 1 else ""}'
        elif diff.seconds >= 60:
            minutes = diff.seconds // 60
            return f'{minutes} minuto{"s" if minutes != 1 else ""}'
        else:
            return 'Ahora mismo'
        
# ============================================================
# EXPORTAR REPORTE
# ============================================================

class ExportReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        gimnasio = request.gimnasio

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # ESTILOS
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        center = Alignment(horizontal="center")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # TÍTULO
        ws.merge_cells('A1:F1')
        ws['A1'] = 'REPORTE DEL GIMNASIO'
        ws['A1'].font = Font(size=18, bold=True)
        ws['A1'].alignment = center

        ws.append([])
        ws.append([f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'])
        ws.append([])

        # DATOS
        if gimnasio:
            miembros = MembresiaAsignada.objects.filter(miembro__gimnasio=gimnasio)
            diarios = UsuarioGymDay.objects.filter(gimnasio=gimnasio)
        else:
            miembros = []
            diarios = []

        total_membresias = sum(m.price for m in miembros)
        total_diarios = sum(d.price for d in diarios)
        total = total_membresias + total_diarios

        # ==========================================================
        # ESTADÍSTICAS
        # ==========================================================
        ws.append(['ESTADÍSTICAS DEL MES'])

        headers = ['Total miembros', 'Ingresos membresías', 'Ingresos diarios', 'Total']
        ws.append(headers)

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        ws.append([
            len(miembros),
            total_membresias,
            total_diarios,
            total
        ])

        # formato dinero
        for col in range(2, 5):
            ws.cell(row=ws.max_row, column=col).number_format = '$#,##0'

        ws.append([])

        # ==========================================================
        # 👤 MIEMBROS
        # ==========================================================
        ws.append(['MIEMBROS'])
        headers = ['Nombre', 'Apellido', 'Membresía', 'Fecha Inicio', 'Precio']
        ws.append(headers)

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        for m in miembros[:50]:
            ws.append([
                m.miembro.name,
                m.miembro.lastname,
                m.membresia.name,
                m.dateInitial.strftime('%d/%m/%Y'),
                m.price
            ])

            row = ws.max_row
            ws.cell(row=row, column=5).number_format = '$#,##0'

        ws.append([])

        # ==========================================================
        # INGRESOS DIARIOS
        # ==========================================================
        ws.append(['INGRESOS DIARIOS'])
        headers = ['Nombre', 'Apellido', 'Fecha', 'Monto']
        ws.append(headers)

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        for d in diarios[:50]:
            ws.append([
                d.name,
                d.lastname,
                d.dateInitial.strftime('%d/%m/%Y'),
                d.price
            ])

            row = ws.max_row
            ws.cell(row=row, column=4).number_format = '$#,##0'

        # ==========================================================
        # AUTO AJUSTE COLUMNAS
        # ==========================================================
        from openpyxl.utils import get_column_letter
        
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    # Ignorar celdas combinadas (MergedCell)
                    if cell.value and not isinstance(cell, type(ws.merged_cells)):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 3 if max_length > 0 else 15

        # ==========================================================
        # DESCARGA
        # ==========================================================
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_gimnasio.xlsx"'

        wb.save(response)
        return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def membership_notifications(request):
    from django.conf import settings
    
    gimnasio = request.gimnasio
    numero_gimnasio = getattr(settings, 'WHATSAPP_NUMBER', '573001234567')  # Default si no está configurado
    
    today = datetime.now().date()
    three_day_later = today + timedelta(days=3)
    
    # Base queryset filtrada por gimnasio
    if gimnasio:
        # Membresías próximas a vencer: solo las que NO han sido marcadas como leídas
        expiring_memberships = MembresiaAsignada.objects.filter(
            miembro__gimnasio=gimnasio,
            dateFinal__gt=today,  # Solo mayores a hoy (no 0 ni negativos)
            dateFinal__lte=three_day_later,
            notified_at__isnull=True  # No marcadas como leídas
        ).select_related('miembro', 'membresia', 'miembro__gimnasio')
        
        # Membresías vencidas: solo las NO leídas
        expired_memberships = MembresiaAsignada.objects.filter(
            miembro__gimnasio=gimnasio,
            dateFinal__lte=today,
            notified_at__isnull=True  # No marcadas como leídas
        ).select_related('miembro', 'membresia', 'miembro__gimnasio')
    else:
        expiring_memberships = MembresiaAsignada.objects.none()
        expired_memberships = MembresiaAsignada.objects.none()

    notifications = []

    for membership in expiring_memberships:
        days_left = (membership.dateFinal - today).days
        member_name = f"{membership.miembro.name} {membership.miembro.lastname}"
        membership_name = membership.membresia.name
        exp_date = membership.dateFinal.strftime('%d/%m/%Y')
        
        # Generar mensaje para WhatsApp
        wa_message = f"Hola%20{member_name},%20tu%20membresía%20({membership_name})%20expira%20el%20{exp_date}%20en%20{days_left}%20días.%20¿Deseas%20renovar%20ahora%20y%20seguir%20entrenando%20con%20nosotros%3F"
        
        # Limpiar teléfono (solo números)
        phone = membership.miembro.phone or ''
        phone_clean = ''.join(filter(str.isdigit, phone))
        
        # Agregar código de país si no lo tiene
        if phone_clean and not phone_clean.startswith('57'):
            phone_clean = '57' + phone_clean
        
        wa_link = f"https://wa.me/{phone_clean}?text={wa_message}" if phone_clean else None
        
        notifications.append({
            'type': 'warning',
            'title': 'Membresía próxima a expirar',
            'message': f'La membresía de {member_name} - {membership_name} expirará en {days_left} días.',
            'date': exp_date,
            'link': f'/dashboard/asignar-membresia/{membership.id}/',
            'whatsapp_link': wa_link,
            'membership_id': membership.id
        })
    
    for membership in expired_memberships:
        member_name = f"{membership.miembro.name} {membership.miembro.lastname}"
        membership_name = membership.membresia.name
        exp_date = membership.dateFinal.strftime('%d/%m/%Y')
        
        # Generar mensaje para WhatsApp
        wa_message = f"Hola%20{member_name},%20tu%20membresía%20({membership_name})%20venció%20el%20{exp_date}.%20¡Te%20esperamos%20de%20vuelta%20para%20que%20renueves%20y%20continúes%20entrenando%21"
        
        # Limpiar teléfono (solo números)
        phone = membership.miembro.phone or ''
        phone_clean = ''.join(filter(str.isdigit, phone))
        
        # Agregar código de país si no lo tiene
        if phone_clean and not phone_clean.startswith('57'):
            phone_clean = '57' + phone_clean
        
        wa_link = f"https://wa.me/{phone_clean}?text={wa_message}" if phone_clean else None
        
        notifications.append({
            'type': 'danger',
            'title': 'Membresía vencida',
            'message': f'La membresía de {member_name} - {membership_name} ya venció.',
            'date': exp_date,
            'link': f'/dashboard/asignar-membresia/{membership.id}/',
            'whatsapp_link': wa_link,
            'membership_id': membership.id
        })
    
    return Response(notifications)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_notifications_read(request):
    """Marca todas las notificaciones como leídas para el gimnasio del usuario."""
    gimnasio = request.gimnasio
    
    if not gimnasio:
        return Response({'status': 'no gimnasio'}, status=status.HTTP_400_BAD_REQUEST)
    
    now = timezone.now()
    
    # Marcar membresías próximas a vencer (próximos 3 días)
    MembresiaAsignada.objects.filter(
        miembro__gimnasio=gimnasio,
        dateFinal__gt=datetime.now().date(),
        dateFinal__lte=datetime.now().date() + timedelta(days=3),
        notified_at__isnull=True
    ).update(notified_at=now)
    
    # Marcar membresías vencidas
    MembresiaAsignada.objects.filter(
        miembro__gimnasio=gimnasio,
        dateFinal__lte=datetime.now().date(),
        notified_at__isnull=True
    ).update(notified_at=now)
    
    return Response({'status': 'ok', 'marked_at': now.isoformat()})